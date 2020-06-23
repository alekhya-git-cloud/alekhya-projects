import sys
import time
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

from PyQt5 import QtCore, QtGui, QtWidgets
import write_to_xlsx_file as wd
import parse_xlsx_file as rd
import openpyxl
import ast

from Score_card import Ui_MainWindow1
Q=20
Qcol=2  
obj=rd.parse_file("CNDS_Test_1.xlsx",Qcol)
list1=obj.get_all_questions()
wd.write_to_xl("output.xlsx",list1,20)

class Ui_Form1(QWidget):
	
	def open_window2(self):
		Window = QtWidgets.QMainWindow()
		ui2 = Ui_MainWindow1()
		ui2.setupUi(Window)
		Window.show()
		app.exec_()
		self.on_clear()
		#self.extbtn.clicked.connect(self.Exam_finished)
	def __init__(self):
		super().__init__()
		self.layout = QVBoxLayout()		

		layout1 = QGridLayout()
		self.setLayout(layout1)


		groupbox = QGroupBox("")
		layout1.addWidget(groupbox)
		#layout1.addWidget(self.s1)
		layout1.setContentsMargins(30,10,10,10)
		#self.layout.setSpacing(10)
		groupbox.setLayout(self.layout)

		self.setupUi()
		self.i=0

	def showTime(self): 
 
		mins, secs = divmod(self.t, 60)
		timeformat = '{:02d}:{:02d}'.format(mins, secs)
		#self.timer.setText(timeformat)
		time.sleep(1)
		self.t -= 1	 
		self.timer.setText(timeformat)

	def setupUi(self):

		self.no_of_q=20
		self.row=1
		
		self.q_no=1
		self.count_marks=0
		self.list2=[]
		self.skip_count=0
		self.setWindowTitle("TS-Programming Aptitude Test")
		self.setGeometry(0, 25, 1380, 700)


		self.logo = QLabel(self)
		self.logo.setGeometry(1150, 15, 250, 50)
		self.logo.setPixmap(QPixmap("thundersoft.png"))

		self.testname = QLabel("C And DS Programming Test",self)
		self.testname.setGeometry(500, 10, 311, 31)
		font = QFont()
		font.setPointSize(15)
		font.setBold(True)
		font.setWeight(75)
		self.testname.setFont(font)
		#self.testname.setAlignment(QtCore.Qt.AlignCenter)
		#self.testname.setObjectName("testname")
		#self.testname.setText("C And DS Programming Test")

		self.timeleft = QLabel("Time Left:",self)
		self.timeleft.setGeometry(1150, 70, 91, 20)
		font = QFont()
		font.setPointSize(12)
		font.setBold(True)
		font.setWeight(75)
		self.timeleft.setFont(font)

		self.t=1200		
		self.timer = QLabel(self)
		self.timer.setGeometry(1250, 70, 91, 20)
		font = QFont()
		font.setPointSize(12)
		font.setBold(True)
		font.setWeight(75)
		self.timer.setFont(font) 
        
        # creating a timer object 
		timer2 = QTimer(self)         
		timer2.timeout.connect(self.showTime)                   
        # update the timer every tenth second 
		timer2.start(100)
		
		self.skipbtn = QPushButton(self)
		self.skipbtn.setGeometry(670, 670, 75, 23)
		self.skipbtn.setText("Skip")

		self.nxtbtn = QPushButton(self)
		self.nxtbtn.setGeometry(770, 670, 75, 23)
		self.nxtbtn.setText("Next")

		self.extbtn = QPushButton(self)
		self.extbtn.setGeometry(1250, 670, 75, 23)
		self.extbtn.setText("Exit")

		#self.submitbtn = QPushButton(self)
		#self.submitbtn.setGeometry(870, 670, 75, 23)
		#self.submitbtn.setText("submit")
		self.path = "output.xlsx"
		self.wb_obj = openpyxl.load_workbook(self.path)
		self.sheet = self.wb_obj.active
		#self.book = openpyxl.Workbook("output.xlsx")
		#self.sheet = self.book.active
		self.attempted_question=0


		self.retranslateUi()
		self.nxtbtn.setEnabled(False)

		self.nxtbtn.clicked.connect(self.nxt_question)
		self.skipbtn.clicked.connect(self.nxt_question)
		#self.extbtn.clicked.connect(self.open_window2)

	def retranslateUi(self):
		self.clear()
		
		if self.row >self.no_of_q:
			self.row=1
		qn_no = self.sheet.cell(row=self.row, column=1).value

		if qn_no <= self.no_of_q:
			ans=self.sheet.cell(row = self.row,column=5).value
			if ans==None:

				self.question = QLabel(self)
				self.question.setGeometry(20,20, 250, 200)
				font = QFont()
				font.setPointSize(12)
				self.question.setFont(font)

	
				ele= self.sheet.cell(row=self.row,column=2).value
				qn_no=self.sheet.cell(row=self.row,column=1).value
				question=str(qn_no)+'.'+str(ele)
				self.question.setText(question)
				self.layout.addWidget(self.question)
				index=1
				list1=self.sheet.cell(row=self.row,column=3).value
				option_list= ast.literal_eval(list1)

				print(option_list,type(option_list))
				for option in option_list:              #print the option
					radiobtn="radiobtn"+str(index)
					btngroup="btngroup"+str(index)
					index=index+1
					print(option)
					self.layout.setContentsMargins(20,50,40,50)
					self.layout.addStretch(1)
					self.radiobtn = QRadioButton(option)
					font = QFont()
					font.setPointSize(12)
					self.radiobtn.setFont(font)
					self.radiobtn.setGeometry(200, 150, 500, 100) 
					self.layout.addWidget(self.radiobtn)
					#self.layout.addWidget(self.s1)
					self.nxtbtn.setEnabled(True)
					self.radiobtn.clicked.connect(self.save_option)

					self.nxtbtn.setEnabled(False)

			else:
				self.clear()
				self.row=self.row+1
				self.retranslateUi()
				
				
		else:
			self.clear()
			self.row=1
			self.retranslateUi()

		
	def nxt_question(self):
		self.row=self.row+1
		if self.row==(self.no_of_q)+1:
			if TypeError:
				pass
			#sys.setrecursionlimit(10**6)
			self.row=1
		if self.attempted_question==Q:

			
			self.nxtbtn.clicked.connect(self.open_window2)
			self.skipbtn.clicked.connect(self.open_window2)
			self.extbtn.clicked.connect(self.open_window2)
		self.retranslateUi()

	def save_option(self):
		self.nxtbtn.setEnabled(True)
		self.radiobtn = self.sender()
		if self.radiobtn.isChecked():
			file1 = open(self.path, 'r+')
			self.data = self.radiobtn.text()
			self.user_data = self.sheet.cell(row=self.row, column=5)
			self.user_data.value = self.data
			self.Ans = self.sheet.cell(row=self.row, column=4)
			self.i=self.i+1
			if self.Ans.value==self.user_data.value:
				self.sheet.cell(row=self.row,column=6).value=1
			else:
				self.sheet.cell(row=self.row,column=6).value=0
			
		self.attempted_question=self.attempted_question+1
		self.wb_obj.save(self.path)
		self.extbtn.clicked.connect(self.open_window2)
	def Exam_finished(self):
		print("all questions have completed")
		dlg = QMessageBox()
		dlg.setIcon(QMessageBox.Question)
		dlg.setText("attempted " +str(self.attempted_question-1)+ "/"+ str(20) +"questions.\n\n Do you want to see Scord Card Details?")
		dlg.exec_()
	def on_clear(self):
		sys.exit()
	def clear(self):
		while self.layout.count():
			child = self.layout.takeAt(0)
			if child.widget():
				child.widget().deleteLater() 


if __name__ == "__main__":
	import sys
	app = QApplication(sys.argv)
	ui = Ui_Form1()
	ui.show()
	sys.exit(app.exec_())
