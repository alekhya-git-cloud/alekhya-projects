import openpyxl
def write_to_xl(file,list1,qn_cnt):
	book=openpyxl.Workbook()
	sheet=book.active
	a=1

	for list2 in list1:
		if a <= qn_cnt:
			c = sheet.cell(row=a, column=1)
			c.value=a       #to add the line number in a file
			col=2
			if list2!=None:
				for ele in list2:
					sheet.cell(row=a,column=col).value=str(ele) #to write a data row wise
					col=col+1
				a=a+1


	book.save(file)
