import openpyxl
import random
import re

class parse_file:
	#to open the input xlsx sheet"
	def __init__(self,file,Qcol):
		self.Qcol=Qcol
		self.ans=3
		self.wb=openpyxl.load_workbook(file)
		self.sheet = self.wb.active
		self.max_row = self.sheet.max_row       #to get the maximum row present in sheet
		self.max_col = self.sheet.max_column    #to get the maximum column present in sheet
		
	#to store all the questions along with its options in a list in the form of:[[Q1,[A,B,C,D]],[Q2,A,B,C,D],...]

	def get_all_questions(self):
		allQ=[]
		for row in range(1,self.max_row+1):
			data=parse_file.get_question_along_with_its_options(self,row)
			allQ.append(data)
		random.shuffle(allQ)
		return allQ

	#to store a single question along with option such as:[Q,[A,B,C,D]]

	def get_question_along_with_its_options(self,row):
		list2=[]
		options=[]
		try:
			for col in range(self.Qcol,self.max_col+1):
				data=self.sheet.cell(row=row,column=col).value
				if col==self.Qcol:		#check whether the column is containing question or option 						 #to get the question in list2
					list2.append(str(data))                         
				elif col==self.ans:
					crct_ans=str(data)
				elif data!=None:
					data=re.search('(?<=\.)(.+)',str(data)).group(0)   #to remove 'A.' ,'B.','C.','D.' from the options
					options.append(str(data))                               #store the options in list(options)
				else:                                                      #if the cell is not containing any answer dont store it
					break

			random.shuffle(options)
			list2.append(options)
			list2.append(crct_ans)
			return list2
		except Exception:
			print("Error in {} row {} column".format(row,col))




			

		


